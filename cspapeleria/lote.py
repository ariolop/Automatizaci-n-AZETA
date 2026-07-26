"""
lote.py — Procesado por lotes de muchos EANs sin bloquear la petición web.

Lanza un hilo que hace un único login y procesa los EANs uno a uno (con una
pequeña pausa para no saturar el WAF). El estado se consulta por sondeo.
"""
import csv
import io
import re
import threading
import time
import uuid

from cs_login import CSSession, CSError
from cs_producto import buscar_producto

_JOBS = {}
_LOCK = threading.Lock()

PAUSA_ENTRE = 0.4  # segundos entre productos


# --- Extracción de EANs ----------------------------------------------------- #
def extraer_eans(texto):
    return list(dict.fromkeys(re.findall(r"\b\d{8,14}\b", texto or "")))


def _col_a_indice(col):
    col = (col or "").strip()
    if not col:
        return None
    if col.isdigit():
        return int(col) - 1
    idx = 0
    for ch in col.upper():
        if "A" <= ch <= "Z":
            idx = idx * 26 + (ord(ch) - 64)
    return idx - 1 if idx else None


def extraer_eans_de_archivo(nombre, contenido, columna=None, fila_inicial=1):
    nombre = (nombre or "").lower()
    idx = _col_a_indice(columna)
    fila0 = max(0, int(fila_inicial or 1) - 1)
    eans = []
    if nombre.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        for i, fila in enumerate(ws.iter_rows(values_only=True)):
            if i < fila0:
                continue
            celdas = fila if idx is None else [fila[idx] if idx < len(fila) else None]
            for c in celdas:
                if c is not None:
                    eans += extraer_eans(str(c))
    else:  # CSV / TXT
        texto = contenido.decode("utf-8-sig", "replace") if isinstance(contenido, bytes) else contenido
        try:
            dialecto = csv.Sniffer().sniff(texto[:2000], delimiters=";,\t")
        except Exception:
            dialecto = csv.excel
            dialecto.delimiter = ";"
        for i, fila in enumerate(csv.reader(io.StringIO(texto), dialecto)):
            if i < fila0:
                continue
            celdas = fila if idx is None else [fila[idx] if idx < len(fila) else ""]
            for c in celdas:
                eans += extraer_eans(str(c))
    return list(dict.fromkeys(eans))


# --- Jobs ------------------------------------------------------------------- #
def crear_job(eans, vigilar=False):
    jid = uuid.uuid4().hex[:12]
    job = {"id": jid, "estado": "en_curso", "total": len(eans),
           "procesados": 0, "items": [], "vigilar": vigilar}
    with _LOCK:
        _JOBS[jid] = job
    threading.Thread(target=_ejecutar, args=(jid, eans, vigilar), daemon=True).start()
    return jid


def _ejecutar(jid, eans, vigilar):
    job = _JOBS[jid]
    try:
        ses = CSSession()
        ses.login()
    except CSError as e:
        job["estado"] = "error"
        job["error"] = str(e)
        return
    import db
    if vigilar:
        db.init_db()
    for ean in eans:
        try:
            f = buscar_producto(ean, sesion=ses)
        except CSError as e:
            f = {"patron": ean, "encontrado": False, "error": str(e)}
        estado = ("desaparecido" if not f.get("encontrado")
                  else "disponible" if f.get("disponible") else "agotado")
        item = {"ean": ean, "estado": estado, "nombre": f.get("nombre"),
                "precio_neto": f.get("precio_neto"), "coste_real": f.get("coste_real"),
                "stock": f.get("stock"),
                "imagen": (f.get("imagenes") or [None])[0]}
        if vigilar and f.get("encontrado"):
            db.add_vigilado(ean, f.get("nombre"))
        with _LOCK:
            job["items"].append(item)
            job["procesados"] += 1
        time.sleep(PAUSA_ENTRE)
    job["estado"] = "terminado"


def obtener_job(jid):
    with _LOCK:
        return dict(_JOBS.get(jid) or {})


def csv_de_job(jid):
    job = obtener_job(jid)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["EAN", "Estado", "Nombre", "PrecioNeto", "Coste", "Stock"])
    for it in job.get("items", []):
        w.writerow([it["ean"], it["estado"], it.get("nombre") or "",
                    it.get("precio_neto") or "", it.get("coste_real") or "",
                    it.get("stock") or ""])
    return buf.getvalue()
